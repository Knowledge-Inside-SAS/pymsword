import os.path
from pymsword.docxcom_template import DocxComTemplate
from pymsword.com_utilities import document_inserter, image_inserter, update_document_toc
import win32com.client
import openpyxl


def _parse_hierarchy_sheet(workbook):
    #get sheet #2
    hier_sheet = workbook.worksheets[1]
    #read header, which is in the first row
    header = []
    for col in range(1, hier_sheet.max_column + 1):
        hdr = hier_sheet.cell(row=1, column=col).value
        if not hdr: break
        header.append(hdr)
    #detect how many level columns are present. Level columns have name "Level X" where X is a number
    num_level_cols = 0
    while num_level_cols < len(header) and header[num_level_cols].startswith("Level "):
        num_level_cols += 1
    #usually we expect 7 of them.
    name2index = {name: i for i, name in enumerate(header)}
    col_type = name2index.get("Type")
    col_rev = name2index.get("Revision")
    col_desc = name2index.get("desc")


    hier_data = [] #list of dictionaries: {"level", "name", "type", "rev", "desc"}
    #level is determined as last non-empty level column in the row
    for row in range(2, hier_sheet.max_row + 1):
        # read the row data
        row_data = [hier_sheet.cell(row=row, column=col).value for col in range(1, len(header) + 1)]
        # detect the level
        level = 0
        name = None
        for i in range(num_level_cols):
            if row_data[i]:
                level = i + 1
                name = row_data[i]
        if level == 0:
            # no level found, empty row
            break
        # create a dictionary for the row
        hier_data.append({
            "level": level,
            "name": name,
            "type": row_data[col_type] if col_type is not None else "n/a",
            "rev": row_data[col_rev] if col_rev is not None else "n/a",
            "desc": row_data[col_desc] if col_desc is not None else "n/a"
        })
    return hier_data

def _parse_history_sheet(workbook):
    history_sheet = workbook.worksheets[0]
    records = []
    row = 3
    while row <= history_sheet.max_row:
        version = history_sheet.cell(row=row, column=1).value
        date_ = history_sheet.cell(row=row, column=3).value
        action = history_sheet.cell(row=row, column=5).value
        desc = history_sheet.cell(row=row, column=7).value
        author = history_sheet.cell(row=row, column=12).value
        if version is None:
            if desc is None:
                # empty row, stop processing
                break
            else:
                # no version, but description, this is a continuation of the previous record
                records[-1]['desc'].append({"name":desc})
        else:
            #new record
            records.append({
                "version": version,
                "date": date_,
                "action": action,
                "desc": [{"name":desc}] if desc else [],
                "author": author
            })
        row += 1
    return records

def parse_excel_data(excel_path:str):
    workbook = openpyxl.load_workbook(excel_path, data_only=True)
    return {"hier":_parse_hierarchy_sheet(workbook),
            "history": _parse_history_sheet(workbook)}


def main():
    template = DocxComTemplate("LAPTOP functional specification template.docx")
    source_dir = "data"


    data = {
        "doc_name": "Functional Specification"
    }
    functions = []
    data['function'] = functions
    #enumerate XSLX files in the source directory
    for filename in sorted(os.listdir(source_dir)):
        if not filename.lower().endswith(".xlsx"): continue
        print("Processing file:", filename)
        func = {}
        functions.append(func)
        funcname = os.path.splitext(filename)[0]
        func['name'] = funcname
        excel_path = os.path.join(source_dir, filename)

        #func['presentation'] = "presentation"
        #func['presentation'] = excel_sheet_inserter(
        #    excel_cache, excel_path, "Presentation"
        #)
        #func['specification'] = excel_sheet_inserter(
        #    excel_cache, excel_path, 2 #sheet index 2 corresponds to the second sheet
        #)
        excel_data = parse_excel_data(excel_path)

        func_rows = []
        for record in excel_data['hier']:
            level = record['level']
            if level > 6: level = 6  # limit to 6 levels, because that's what we have in the template.
            func_rows.append({f"row{level}": [record]})
        func['tabrow'] = func_rows

        func['history_row'] = excel_data['history']


        description_file = os.path.join(source_dir, funcname + ".rtf")
        if os.path.exists(description_file):
            func['description'] = document_inserter(description_file, set_style="Normal")

        diagram_file = os.path.join(source_dir, funcname + ".svg")
        if os.path.exists(diagram_file):
            func['diagram'] = image_inserter(diagram_file)


    template.generate(data, "LAPTOP functional specification result.docx",
                      postprocess=update_document_toc)
    os.startfile("LAPTOP functional specification result.docx")
if __name__ == "__main__":
    main()
    # Make sure to close the Excel application when done
